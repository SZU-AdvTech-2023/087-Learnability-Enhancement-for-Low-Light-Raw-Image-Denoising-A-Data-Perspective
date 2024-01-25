import toupcam
import numpy as np
from PyQt5.QtGui import QPixmap, QImage
from PIL import Image
import tifffile as tiff
import os
import glob
import threading
import time
import re
import openpyxl
import json
from scipy.stats import linregress

class App:
    def __init__(self):
        self.hcam = None
        self.buf = None
        self.total = 0
        self.width =  0
        self.height = 0
        self.ExpoTime = None
        self.ExpoAGain = None
        self.BLE = None
        self.FPN = None
        self.snap = None
        self.BLEpath = None
        self.FPNpath = None
        self.snappath = None
        self.image_stack = [] #用于融合图片

# the vast majority of callbacks come from toupcam.dll/so/dylib internal threads
    @staticmethod
    def cameraCallback(nEvent, ctx):
        if nEvent == toupcam.TOUPCAM_EVENT_IMAGE:
            ctx.CameraCallback(nEvent)

    def CameraCallback(self, nEvent):
        if nEvent == toupcam.TOUPCAM_EVENT_IMAGE:
            try:
                ExpoTime = self.hcam.get_ExpoTime()  #获取当前图片的曝光时间和增益系数
                ExpoAGain = self.hcam.get_ExpoAGain()
                print(f"{self.total+1}:{ExpoTime}-{ ExpoAGain }")
                
                self.hcam.PullImageV3(self.buf, 0, 16, 0, None)
                self.total += 1
                current_thread_id = threading.current_thread().ident
                print(f'pull image ok, total = {self.total}   {current_thread_id}')
                
            except toupcam.HRESULTException as ex:
                print('pull image failed, hr=0x{:x}'.format(ex.hr & 0xffffffff))

            else:
                if self.BLE == 1:
                    self.onBLEprocess() 
                elif self.FPN == 1:
                    self.onFPNprocess()
                elif self.snap == 1:
                    self.onSnap()
        else:
            print('event callback: {}'.format(nEvent))

    def onSnap(self):
        ExpoTime = self.hcam.get_ExpoTime()  #获取当前图片的曝光时间和增益系数
        ExpoAGain = self.hcam.get_ExpoAGain()
        if ExpoTime == self.ExpoTime and ExpoAGain == self.ExpoAGain:  #当前参数和之前参数不同时，暂时不响应这个事件，防止读取之前参数设置拍摄的图片
            image = QImage(self.buf, self.width, self.height, QImage.Format_Grayscale16)  
            ptr = image.constBits()
            image_arr = np.frombuffer(ptr.asstring(2048*2048*2), dtype=np.uint16).reshape(self.height, -1)* 16 
            self.image_stack.append(image_arr)
            if len(self.image_stack)==10: #十帧融合图片
                image_ave = np.mean(self.image_stack, axis=0).astype(np.uint16)
                os.makedirs(self.FPNpath, exist_ok=True)
                file_name = f"{ExpoTime}_{ExpoAGain}.tif"
                tiff.imwrite(file_name,image_ave)
                self.image_stack.clear() #清空列表用于下一次融合
                self.snap = 0 #终止处理

        else:  #下一次再读取缓存，保证设置刷新
            self.ExpoTime = ExpoTime
            self.ExpoAGain = ExpoAGain

    def onBLEprocess(self):
        ExpoTime = self.hcam.get_ExpoTime()  #获取当前图片的曝光时间和增益系数
        ExpoAGain = self.hcam.get_ExpoAGain()
        print(f"{self.total}:{ExpoTime}-{ ExpoAGain }")
        if ExpoTime == self.ExpoTime and ExpoAGain == self.ExpoAGain:
            image = QImage(self.buf, self.width, self.height, QImage.Format_Grayscale16)  
            ptr = image.constBits()
            image_arr = np.frombuffer(ptr.asstring(2048*2048*2), dtype=np.uint16).reshape(self.height, -1)* 16

            os.makedirs(self.BLEpath, exist_ok=True)
            file_name = f"{self.BLEpath}/{ExpoTime}_{ExpoAGain}.tif"
            tiff.imwrite(file_name ,image_arr)
            
            #系数变化逻辑
            ExpoAGain = ExpoAGain + 10
            if ExpoAGain == 810:
                ExpoTime = ExpoTime + 100000
                ExpoAGain = 100
            if  ExpoTime > 1000000:
                self.BLE = 0 #终止处理
            
            self.hcam.put_ExpoTime(ExpoTime)
            self.hcam.put_ExpoAGain(ExpoAGain)

        else:  #下一次再读取缓存，保证设置刷新
            self.ExpoTime = ExpoTime
            self.ExpoAGain = ExpoAGain

    def onFPNprocess(self):
        ExpoTime = self.hcam.get_ExpoTime()  #获取当前图片的曝光时间和增益系数
        ExpoAGain = self.hcam.get_ExpoAGain()
        print(f"{self.total}:{ExpoTime}-{ ExpoAGain }")
        if ExpoAGain == self.ExpoAGain:
            image = QImage(self.buf, self.width, self.height, QImage.Format_Grayscale16)  
            ptr = image.constBits()
            image_arr = np.frombuffer(ptr.asstring(2048*2048*2), dtype=np.uint16).reshape(self.height, -1)* 16 
            self.image_stack.append(image_arr)
            if len(self.image_stack)==10: #十帧融合图片
                image_ave = np.mean(self.image_stack, axis=0).astype(np.uint16)
                os.makedirs(self.FPNpath, exist_ok=True)
                file_name = f"{self.FPNpath}/{ExpoTime}_{ExpoAGain}.tif"
                tiff.imwrite(file_name,image_ave)

                self.image_stack.clear() #清空列表用于下一次融合
                ExpoAGain = ExpoAGain + 10
                if ExpoAGain == 810:
                    ExpoAGain = 100
                    self.FPN = 0 #终止循环
                self.hcam.put_ExpoAGain(ExpoAGain)

        else:  #下一次再读取缓存，保证设置刷新
            self.ExpoAGain = ExpoAGain

        

    def BLEpreprocess(self, BLEpath = "camera/autoBLE"):
        self.BLEpath =  BLEpath
        self.hcam.put_ExpoTime(100000)  #初始曝光时间
        self.hcam.put_ExpoAGain(100)       #初始增益系数
        self.ExpoTime = self.hcam.get_ExpoTime()  #获取当前图片的曝光时间和增益系数
        self.ExpoAGain = self.hcam.get_ExpoAGain()
        print(f"initial:{self.ExpoTime}-{self.ExpoAGain }")
        time.sleep(1)  #更改设置后，等待缓冲刷新
        #获取图片到指定文件夹
        self.BLE = 1;#处理黑电平模式
        while self.BLE == 1:
            time.sleep(1)
        self.BLE = None #关闭处理
        #处理图片数据存到Excel
        process_BLEimages(BLEpath)
        fit_linear_model_to_data_and_save('output.xlsx')
    
    def FPNpreprocess(self, FPNpath = "camera/autoFPN", ExpoTime = 100000):
        self.FPNpath = FPNpath
        self.hcam.put_ExpoTime(ExpoTime)  #初始曝光时间，100ms
        self.hcam.put_ExpoAGain(100)       #初始增益系数
        self.ExpoTime = self.hcam.get_ExpoTime()  #获取当前图片的曝光时间和增益系数
        self.ExpoAGain = self.hcam.get_ExpoAGain()
        print(f"initial:{self.ExpoTime}-{self.ExpoAGain }")
        time.sleep(1) #更改设置后，等待缓冲刷新
        #获取图片到指定文件夹
        self.FPN = 0;#处理FPN模式
        while self.FPN==1:
            time.sleep(1)
        self.FPN = None #关闭处理
        #对FPN进行线性拟合，并获得两个参数矩阵
        iso_params = read_iso_params_from_json('iso_params.json')  #BLE参数
        iso_values, image_stack = process_FPNimages(FPNpath,100000,iso_params) #获取剔除BLE的iso和相应图像矩阵
        FPN_linear_fit(iso_values, image_stack)  #线性拟合并存储两个矩阵
    
    def Snap(self, ExpoTime = 100000, ExpoGain = 100):
        self.hcam.put_ExpoTime(ExpoTime)  #初始曝光时间，100ms
        self.hcam.put_ExpoAGain(ExpoGain)       #初始增益系数
        self.ExpoTime = self.hcam.get_ExpoTime()  #获取当前图片的曝光时间和增益系数
        self.ExpoAGain = self.hcam.get_ExpoAGain()
        print(f"initial:{self.ExpoTime}-{self.ExpoAGain }")
        time.sleep(1)  #更改设置后，等待缓冲刷新
        #获取图片到指定文件夹
        self.snap = 1;#处理单拍模式
        while self.snap==1:
            time.sleep(1)
        self.snap = None #关闭处理

    def run(self):
        a = toupcam.Toupcam.EnumV2()
        if len(a) > 0:
            print('{}: flag = {:#x}, preview = {}, still = {}'.format(a[0].displayname, a[0].model.flag, a[0].model.preview, a[0].model.still))
            for r in a[0].model.res:
                print('\t = [{} x {}]'.format(r.width, r.height))   #输出分辨率选项
            self.hcam = toupcam.Toupcam.Open(a[0].id)   #打开第一个相机
            if self.hcam:
                try:
                    self.width, self.height = self.hcam.get_Size()
                    bufsize = self.width * self.height * 2
                    self.hcam.put_Option(toupcam.TOUPCAM_OPTION_BYTEORDER, 0)   #Qimage use RGB byte order
                    self.hcam.put_Option(toupcam.TOUPCAM_OPTION_RAW, 0)   #Qimage use RGB byte order
                    self.hcam.put_Option(toupcam.TOUPCAM_OPTION_BITDEPTH, 1)    #channel depth 12bit
                    self.hcam.put_Option(toupcam.TOUPCAM_OPTION_RGB, 4)         #mode 16 gray
                    self.hcam.put_Option(toupcam.TOUPCAM_OPTION_UPSIDE_DOWN, 0)
                    self.hcam.put_Option(toupcam.TOUPCAM_OPTION_MULTITHREAD,0)
                    self.hcam.put_AutoExpoEnable(0)   #自动曝光关
                    self.hcam.put_ExpoTime(100000)  #
                    self.hcam.put_ExpoAGain(100)      

                    print('image size: {} x {}, bufsize = {}'.format(self.width, self.height, bufsize))
                    self.buf = bytes(bufsize)
                    if self.buf:
                        try:
                            self.hcam.StartPullModeWithCallback(self.cameraCallback, self)
                        except toupcam.HRESULTException as ex:
                            print('failed to start camera, hr=0x{:x}'.format(ex.hr & 0xffffffff))

                    #self.BLEpreprocess("camera/autoBLE")#预处理获取各个ISO下的BLE参数
                    #self.FPNpreprocess("camera/autoFPN")#预处理获取固定曝光时间的FPN图片，剔除BLE，并拟合得到两个矩阵参数
                    #self.Snap(200000,200) #单拍
                    #input('press ENTER to exit')
       
                finally:
                    self.hcam.Close()   #关闭相机
                    self.hcam = None
                    self.buf = None
            else:
                print('failed to open camera')
        else:
            print('no camera found')

    def DSC(self, image_path,iso_value, exposure_time):
        iso_params = read_iso_params_from_json('iso_params.json')
        K_matrix, B_matrix = load_matrices()

        img = Image.open(image_path)
        img_matrix = np.array(img)
    
        ble = calculate_pixel_value(iso_value, exposure_time, iso_params)
        fpn = iso_value*K_matrix + B_matrix
        img_matrix = img_matrix - ble - fpn
        # 将小于0的值替换为0
        img_matrix[img_matrix < 0] = 0
        # 转换为16位整数
        img_matrix = img_matrix.astype(np.int16)

        file_name = f"DSC_{image_path}"
        tiff.imwrite(file_name,img_matrix)

    def shadow(self,iso_value, exposure_time):
        iso_params = read_iso_params_from_json('iso_params.json')
        K_matrix, B_matrix = load_matrices()

        ble = calculate_pixel_value(iso_value, exposure_time, iso_params)
        fpn = iso_value*K_matrix + B_matrix
        img_matrix = ble + fpn
        # 将小于0的值替换为0
        img_matrix[img_matrix < 0] = 0
        # 转换为16位整数
        img_matrix = img_matrix.astype(np.int16)

        file_name = f"shadow_{exposure_time}_{iso_value}.tif"
        tiff.imwrite(file_name,img_matrix)

#utils
def calculate_average_pixel_value(image_path):
    """ 计算图片像素的平均值 """
    with Image.open(image_path) as img:
        img_array = np.array(img)
        return np.mean(img_array)

def get_coordinates_from_filename(filename):
    """ 从文件名中提取坐标 """
    match = re.match(r'(\d+)_(\d+)', filename)
    if match:
        return int(match.group(1)), int(match.group(2))
    return None, None

    #计算相应曝光时间和ISO图片的BLE，存为表格
def process_BLEimages(folder_path):
    """ 处理指定文件夹中的所有图片 """
    wb = openpyxl.Workbook()
    ws = wb.active

    # 收集所有的曝光时间和增益系数
    exposures = set()
    gains = set()
    for filename in os.listdir(folder_path):
        if filename.lower().endswith(('.png', '.jpg', '.jpeg', '.tif', '.tiff', '.bmp')):
            exposure, gain = get_coordinates_from_filename(filename)
            if exposure is not None and gain is not None:
                exposures.add(exposure)
                gains.add(gain)

    # 排序
    sorted_exposures = sorted(exposures)
    sorted_gains = sorted(gains)

    # 在Excel中创建标题行和列
    for i, exposure in enumerate(sorted_exposures, start=2):
        ws.cell(row=1, column=i, value=exposure)
    for i, gain in enumerate(sorted_gains, start=2):
        ws.cell(row=i, column=1, value=gain)

    # 填充数据
    for filename in os.listdir(folder_path):
        if filename.lower().endswith(('.png', '.jpg', '.jpeg', '.tif', '.tiff', '.bmp')):
            exposure, gain = get_coordinates_from_filename(filename)
            if exposure is not None and gain is not None:
                avg_pixel_value = calculate_average_pixel_value(os.path.join(folder_path, filename))
                exposure_index = sorted_exposures.index(exposure) + 2
                gain_index = sorted_gains.index(gain) + 2
                ws.cell(row=gain_index, column=exposure_index, value=avg_pixel_value)

    wb.save('output.xlsx')

#根据Excel，拟合BLE每个ISO下的线性函数，并存成Excel和json
def fit_linear_model_to_data_and_save(excel_file = 'output.xlsx',json_file = 'iso_params.json'):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    exposure_times = [cell.value for cell in ws[1] if cell.value is not None]

    # 准备存储斜率和截距
    last_column = ws.max_column
    ws.cell(row=1, column=last_column + 1, value="Slope (k)")
    ws.cell(row=1, column=last_column + 2, value="Intercept (b)")

    iso_params = {}

    for row_index, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1), start=2):
        gain = row[0].value
        if gain is not None:
            pixel_values = [ws.cell(row=row_index, column=i+2).value for i in range(len(exposure_times))]
            slope, intercept, _, _, _ = linregress(exposure_times, pixel_values)
            ws.cell(row=row_index, column=last_column + 1, value=slope)
            ws.cell(row=row_index, column=last_column + 2, value=intercept)
            iso_params[gain] = {'slope': slope, 'intercept': intercept}

    # 保存修改后的Excel文件
    wb.save('fitted_output.xlsx')

    # 保存为JSON文件
    with open(json_file, 'w') as f:
        json.dump(iso_params, f, indent=4, sort_keys=True)

#读json存储的BLE参数线性拟合参数
def read_iso_params_from_json(json_file = 'iso_params.json'):
    with open(json_file, 'r') as f:
        iso_params = json.load(f)
    return iso_params

#计算对应曝光时间和ISO下的BLE
def calculate_pixel_value(iso, exposure_time, iso_params):
    iso = str(iso)  #转为字符串去索引
    if iso in iso_params:
        slope = iso_params[iso]['slope']
        intercept = iso_params[iso]['intercept']
        return slope * exposure_time + intercept
    else:
        return None

def process_FPNimages(folder_path, exposure_time, iso_params):
    images = glob.glob(os.path.join(folder_path, '*.tif'))
    iso_values = []
    image_stack = []

    for image_path in images:
        iso_value = int(image_path.split('_')[-1].split('.')[0])
        img = Image.open(image_path)
        img_matrix = np.array(img)

        # 减去基线曝光
        ble = calculate_pixel_value(iso_value, exposure_time, iso_params)
        if ble is not None:
            img_matrix = img_matrix - ble
        else:
            print(f"{iso_value}ISO值对应的BLE未计算")

        iso_values.append(iso_value)
        image_stack.append(img_matrix)

    return np.array(iso_values), np.stack(image_stack, axis=-1)

def FPN_linear_fit(iso_values, image_stack):
    x = iso_values
    x_mean = x.mean()
    # 添加一个新维度，使形状变为 (2048, 2048, 1)
    y_mean = image_stack.mean(axis=-1)[:, :, np.newaxis]
    ss_xy = np.tensordot((x - x_mean), (image_stack - y_mean), axes=([0], [2]))
    ss_xx = np.sum((x - x_mean) ** 2)

    K_matrix = ss_xy / ss_xx
    # 扩展 K_matrix 的形状到 (2048, 2048, 1) 以匹配 y_mean
    K_matrix_expanded = K_matrix[:, :, np.newaxis]
    B_matrix = y_mean - K_matrix_expanded  * x_mean
    B_matrix = B_matrix.squeeze()
    save_matrices(K_matrix,B_matrix)

    return K_matrix, B_matrix

def save_matrices(K_matrix, B_matrix, k_matrix_path = 'K_matrix.npy', b_matrix_path = 'B_matrix.npy'):
    np.save( k_matrix_path, K_matrix)
    np.save( b_matrix_path, B_matrix)

def load_matrices(k_matrix_path = 'K_matrix.npy', b_matrix_path = 'B_matrix.npy'):
    K_matrix = np.load(k_matrix_path)
    B_matrix = np.load(b_matrix_path)
    return K_matrix, B_matrix

if __name__ == "__main__":
    app = App()
    #app.run()
    #app.DSC("200000_200.tif",200,200000)
    app.DSC("10000_600.tif",600,10000)
    app.DSC("20000_300.tif",300,20000)
    app.DSC("30000_200.tif",200,30000)
    #app.shadow(200,200000)